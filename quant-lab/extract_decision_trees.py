"""
Extract ALL decision trees, playbooks, and conditional mappings
from the Holy Grail Excel into structured JSON for ML training.
"""
import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx", read_only=True)

decision_trees = {}

# 1. DECISION TREE - WEEKLY CLOSE (full extraction)
ws = wb["DECISION TREE - WEEKLY CLOSE"]
tree_data = []
for row in ws.iter_rows(values_only=True):
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        tree_data.append(non_empty)

decision_trees["weekly_close_decision_tree"] = {
    "source": "Holy Grail Excel - DECISION TREE - WEEKLY CLOSE",
    "description": "EUR/USD Deterministic Model - Navigating Daily Delivery to Weekly Close",
    "data": "Based on 313 Monday sessions, 281 validated weeks (Jan 2020 - Dec 2025)",
    "rows": tree_data
}

# 2. PHASE 5 - WILM ILM VELOCITY ANALYSIS
ws2 = wb["PHASE 5 - WILM ILM VELOCITY ANALYSIS"]
phase5_data = []
for row in ws2.iter_rows(values_only=True):
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        phase5_data.append(non_empty)

decision_trees["phase5_ilm_velocity"] = {
    "source": "Holy Grail Excel - PHASE 5 WILM ILM VELOCITY ANALYSIS",
    "rows": phase5_data
}

# 3. PHASE 6 - SESSION PROFILE SYNTHESIS
ws3 = wb["PHASE 6 - SESSION PROFILE SYNTHESIS"]
phase6_data = []
for row in ws3.iter_rows(values_only=True):
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        phase6_data.append(non_empty)

decision_trees["phase6_session_playbooks"] = {
    "source": "Holy Grail Excel - PHASE 6 SESSION PROFILE SYNTHESIS",
    "rows": phase6_data
}

# 4. PHASE 4 sheets
for sheet_name in ["PHASE 4 - MONTHLY RANGE RECONNAISSANCE", "PHASE 4A - MONTHLY DATASET",
                    "PHASE 4B - RANGE WINDOW TESTING", "PHASE 4C - GROUP B ANALYSIS",
                    "PHASE 4 - TEMPORAL DELIVERY MAPPING"]:
    if sheet_name in wb.sheetnames:
        ws4 = wb[sheet_name]
        rows = []
        for row in ws4.iter_rows(values_only=True):
            non_empty = [str(v) for v in row if v is not None]
            if non_empty:
                rows.append(non_empty)
        decision_trees[sheet_name.lower().replace(" ", "_").replace("-", "_")] = {
            "source": f"Holy Grail Excel - {sheet_name}",
            "rows": rows
        }

# 5. Validation Checklist
ws5 = wb["Validation Checklist"]
vc_data = []
for row in ws5.iter_rows(values_only=True):
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        vc_data.append(non_empty)
decision_trees["validation_checklist"] = {
    "source": "Holy Grail Excel - Validation Checklist",
    "rows": vc_data
}

# 6. Top 10 Claims
ws6 = wb["Top 10 Claims - Testing Framework"]
tc_data = []
for row in ws6.iter_rows(values_only=True):
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        tc_data.append(non_empty)
decision_trees["top_10_claims"] = {
    "source": "Holy Grail Excel - Top 10 Claims Testing Framework",
    "rows": tc_data
}

# 7. Hit Rate Analysis Framework
ws7 = wb["Hit Rate Analysis Framework"]
hr_data = []
for row in ws7.iter_rows(values_only=True):
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        hr_data.append(non_empty)
decision_trees["hit_rate_analysis_framework"] = {
    "source": "Holy Grail Excel - Hit Rate Analysis Framework",
    "rows": hr_data
}

wb.close()

# Save
output_path = r"C:\Users\wifik\Desktop\projects\larger-lab\quant-lab\data\holy_grail_extracted\decision_trees_playbooks.json"
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(decision_trees, f, indent=2, ensure_ascii=False)

print(f"Extracted {len(decision_trees)} decision tree / playbook sections")
print(f"Saved to: {output_path}")
for key in decision_trees:
    n_rows = len(decision_trees[key].get("rows", []))
    print(f"  {key}: {n_rows} rows")

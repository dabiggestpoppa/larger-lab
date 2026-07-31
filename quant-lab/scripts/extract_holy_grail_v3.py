"""
CEREBUS Holy Grail Excel Extractor v3
======================================
Handles sparse sheets with inflated max_row/max_column.
Uses early termination when consecutive empty rows are found.
"""

import openpyxl
import json
import os
import csv
import re
from datetime import datetime, date
from typing import Any, Dict, List, Optional

EXCEL_PATH = r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx"
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "reports", "predecessor", "extracted"
)
PRICE_DATA_DIR = os.path.join(OUTPUT_DIR, "price_data")
SHEET_DATA_DIR = os.path.join(OUTPUT_DIR, "sheet_data")
ONTOLOGY_DIR = os.path.join(OUTPUT_DIR, "ontology_stats")

MAX_EMPTY_ROWS = 5  # Stop after this many consecutive empty rows
MAX_COLS = 30       # Max columns to read


def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return str(val).strip()


def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace(",", "").replace("%", "")
        return float(s)
    except (ValueError, TypeError):
        return None


def is_price_sheet(sheet_name, ws):
    """Check if sheet contains OHLCV price data."""
    price_patterns = [
        r'_(H1|H4|M15|M5|D1|W1|MN)_',
        r'_RAW_DATA$', r'_Daily_', r'_Weekly_',
        r'ETH_(M15|H1|DATA)', r'OILUSD_(H1|H4)',
        r'DAILY DELIVERY NAVIGATION', r'ETH_RANGE_EXPLORATION',
        r'PHASE 4C - THRESHOLD TEST RESULTS',
    ]
    for p in price_patterns:
        if re.search(p, sheet_name, re.IGNORECASE):
            return True
    
    # Check first cell for MT5 export format
    first_cell = ""
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True):
        first_cell = str(row[0]) if row and row[0] else ""
    if '<DATE>' in first_cell.upper() or '_x0009_' in first_cell:
        return True
    return False


def read_sheet_data(ws):
    """
    Read actual data from a sheet, handling sparse data.
    Returns (headers, data_rows, actual_row_count).
    Stops after MAX_EMPTY_ROWS consecutive empty rows.
    """
    headers = []
    data_rows = []
    header_row = 1
    empty_count = 0
    row_num = 0
    
    for row in ws.iter_rows(min_row=1, max_col=MAX_COLS, values_only=True):
        row_num += 1
        
        # Check if row has any data
        has_data = any(v is not None for v in row)
        
        if not has_data:
            empty_count += 1
            if empty_count >= MAX_EMPTY_ROWS and data_rows:
                break
            continue
        
        empty_count = 0
        
        if not headers:
            # First non-empty row = headers
            headers = [safe_str(v) if v else f"col_{j+1}" for j, v in enumerate(row)]
            header_row = row_num
            continue
        
        # Data row
        row_data = {}
        for j, val in enumerate(row):
            if val is not None and j < len(headers):
                fval = safe_float(val)
                row_data[headers[j]] = fval if fval is not None else safe_str(val)
        if row_data:
            data_rows.append(row_data)
    
    return headers, data_rows, row_num


def extract_price_csv(ws, sheet_name):
    """Extract OHLCV price data to CSV."""
    safe_name = re.sub(r'[^\w\-]', '_', sheet_name)[:80]
    csv_path = os.path.join(PRICE_DATA_DIR, f"{safe_name}.csv")
    
    # Read first cell to detect format
    first_cell = ""
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True):
        first_cell = str(row[0]) if row and row[0] else ""
    
    is_mt5_export = '<DATE>' in first_cell.upper() or '_x0009_' in first_cell
    
    rows_written = 0
    empty_count = 0
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["date", "time", "open", "high", "low", "close", "tickvol", "spread"])
        
        header_seen = False
        for row in ws.iter_rows(min_row=1, max_col=10, values_only=True):
            if not any(v is not None for v in row):
                empty_count += 1
                if empty_count >= 5:
                    break
                continue
            empty_count = 0
            
            if not header_seen:
                header_seen = True
                continue
            
            if is_mt5_export and row[0] and '\t' in str(row[0]):
                parts = str(row[0]).split('\t')
                if len(parts) >= 6:
                    try:
                        writer.writerow([
                            parts[0], parts[1] if len(parts) > 1 else "",
                            float(parts[2]), float(parts[3]), float(parts[4]), float(parts[5]),
                            float(parts[6]) if len(parts) > 6 else 0,
                            float(parts[8]) if len(parts) > 8 else 0
                        ])
                        rows_written += 1
                    except (ValueError, IndexError):
                        pass
            elif not is_mt5_export:
                # Column-based format
                date_val = str(row[0]) if row[0] else ""
                time_val = str(row[1]) if len(row) > 1 and row[1] else ""
                o = safe_float(row[2]) if len(row) > 2 else None
                h = safe_float(row[3]) if len(row) > 3 else None
                l = safe_float(row[4]) if len(row) > 4 else None
                c = safe_float(row[5]) if len(row) > 5 else None
                if o and h and l and c:
                    writer.writerow([date_val, time_val, o, h, l, c,
                                     safe_float(row[6]) or 0 if len(row) > 6 else 0,
                                     safe_float(row[8]) or 0 if len(row) > 8 else 0])
                    rows_written += 1
    
    return csv_path if rows_written > 0 else None


def extract_stats(ws):
    """Extract key numerical stats from sheet."""
    stats = {}
    empty_count = 0
    row_num = 0
    
    for row in ws.iter_rows(min_row=1, max_col=MAX_COLS, values_only=True):
        row_num += 1
        
        if not any(v is not None for v in row):
            empty_count += 1
            if empty_count >= MAX_EMPTY_ROWS:
                break
            continue
        empty_count = 0
        
        for col_idx, val in enumerate(row):
            if val is None:
                continue
            
            val_str = safe_str(val)
            
            # Percentage pattern
            pct_match = re.search(r'(\d+\.?\d*)%', val_str)
            if pct_match:
                pct_val = float(pct_match.group(1))
                # Get context from preceding columns
                ctx_parts = [safe_str(row[c]) for c in range(col_idx) if row[c] and len(safe_str(row[c])) < 80]
                ctx = " | ".join(ctx_parts[-3:])
                key = re.sub(r'[^\w\s]', '', ctx)[:50].strip().replace(' ', '_') or f"pct_r{row_num}_c{col_idx}"
                stats[key] = {"value": pct_val, "type": "percentage", "context": ctx, "row": row_num}
            
            # Ratio pattern (0.XXX floats)
            if isinstance(val, float) and 0 < val < 1:
                ctx_parts = [safe_str(row[c]) for c in range(col_idx) if row[c] and len(safe_str(row[c])) < 80]
                ctx = " | ".join(ctx_parts[-3:])
                key = re.sub(r'[^\w\s]', '', ctx)[:50].strip().replace(' ', '_') or f"ratio_r{row_num}_c{col_idx}"
                stats[key] = {"value": val, "type": "ratio", "context": ctx, "row": row_num}
    
    return stats


def main():
    print("=" * 70)
    print("CEREBUS Holy Grail Excel Extractor v3")
    print("=" * 70)
    
    os.makedirs(SHEET_DATA_DIR, exist_ok=True)
    os.makedirs(PRICE_DATA_DIR, exist_ok=True)
    os.makedirs(ONTOLOGY_DIR, exist_ok=True)
    
    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Total sheets: {len(wb.sheetnames)}\n")
    
    summary = {
        "extraction_time": datetime.now().isoformat(),
        "source_file": EXCEL_PATH,
        "total_sheets": len(wb.sheetnames),
        "sheets": {},
        "price_data_files": [],
        "total_data_points": 0,
        "total_stats_extracted": 0,
    }
    
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        price = is_price_sheet(sheet_name, ws)
        stype = "price" if price else "analysis"
        
        print(f"[{i+1:3d}/{len(wb.sheetnames)}] {sheet_name[:50]:50s} [{stype}]", end="", flush=True)
        
        info = {"index": i, "type": stype, "rows": ws.max_row or 0, "cols": ws.max_column or 0}
        
        try:
            if price:
                csv_path = extract_price_csv(ws, sheet_name)
                if csv_path:
                    n = sum(1 for _ in open(csv_path, encoding='utf-8')) - 1
                    info["csv_file"] = os.path.basename(csv_path)
                    info["price_rows"] = n
                    summary["price_data_files"].append({"sheet": sheet_name, "file": os.path.basename(csv_path), "rows": n})
                    summary["total_data_points"] += n
                    print(f" -> {n:,} price rows")
                else:
                    print(" -> No price data, reading as table...")
                    headers, data_rows, actual_rows = read_sheet_data(ws)
                    jp = os.path.join(SHEET_DATA_DIR, f"sheet_{i:03d}.json")
                    with open(jp, 'w', encoding='utf-8') as f:
                        json.dump({"sheet_name": sheet_name, "headers": headers, "row_count": len(data_rows), "data": data_rows}, f, indent=2, default=str)
                    info["json_file"] = f"sheet_{i:03d}.json"
                    info["data_rows"] = len(data_rows)
                    summary["total_data_points"] += len(data_rows)
                    print(f"   -> {len(data_rows)} rows")
            else:
                headers, data_rows, actual_rows = read_sheet_data(ws)
                jp = os.path.join(SHEET_DATA_DIR, f"sheet_{i:03d}.json")
                with open(jp, 'w', encoding='utf-8') as f:
                    json.dump({"sheet_name": sheet_name, "headers": headers, "row_count": len(data_rows), "data": data_rows}, f, indent=2, default=str)
                
                info["json_file"] = f"sheet_{i:03d}.json"
                info["data_rows"] = len(data_rows)
                info["headers"] = headers[:12]
                summary["total_data_points"] += len(data_rows)
                
                stats = extract_stats(ws)
                if stats:
                    sp = os.path.join(ONTOLOGY_DIR, f"stats_{i:03d}.json")
                    with open(sp, 'w', encoding='utf-8') as f:
                        json.dump(stats, f, indent=2, default=str)
                    info["stats_count"] = len(stats)
                    summary["total_stats_extracted"] += len(stats)
                
                print(f" -> {len(data_rows)} rows, {len(stats)} stats")
        
        except Exception as e:
            info["error"] = str(e)
            print(f" -> ERROR: {e}")
        
        summary["sheets"][sheet_name] = info
    
    wb.close()
    
    # Save summary
    sp = os.path.join(OUTPUT_DIR, "summary.json")
    with open(sp, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, default=str)
    
    # Generate log
    lp = os.path.join(OUTPUT_DIR, "extraction_log.md")
    with open(lp, 'w', encoding='utf-8') as f:
        f.write("# CEREBUS Holy Grail — Extraction Report\n\n")
        f.write(f"**Extracted:** {summary['extraction_time']}\n")
        f.write(f"**Source:** `{EXCEL_PATH}`\n")
        f.write(f"**Total Sheets:** {summary['total_sheets']}\n")
        f.write(f"**Total Data Points:** {summary['total_data_points']:,}\n")
        f.write(f"**Total Stats Extracted:** {summary['total_stats_extracted']:,}\n\n")
        
        f.write("## Price Data Files\n\n")
        f.write("| Sheet | File | Rows |\n|-------|------|------|\n")
        for pf in summary["price_data_files"]:
            f.write(f"| {pf['sheet'][:50]} | {pf['file']} | {pf['rows']:,} |\n")
        
        f.write("\n## All Sheets\n\n")
        f.write("| # | Sheet | Type | Data Rows | Stats |\n")
        f.write("|---|-------|------|-----------|-------|\n")
        for name, inf in summary["sheets"].items():
            dr = inf.get("data_rows", inf.get("price_rows", "-"))
            sc = inf.get("stats_count", "-")
            f.write(f"| {inf['index']+1} | {name[:45]:45s} | {inf['type']:8s} | {dr} | {sc} |\n")
    
    print(f"\n{'='*70}")
    print(f"DONE: {summary['total_data_points']:,} data points, {summary['total_stats_extracted']:,} stats")
    print(f"Price files: {len(summary['price_data_files'])}")
    print(f"Summary: {sp}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()

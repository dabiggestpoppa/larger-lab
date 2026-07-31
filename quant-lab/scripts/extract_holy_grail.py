"""
CEREBUS Holy Grail Excel Extractor
===================================
Extracts EVERY stat from EVERY sheet in the predecessor Excel file.
Handles mixed formats: structured tables, embedded analysis, raw price data.
Outputs structured JSON for ontology ingestion.

Usage:
    python quant-lab/scripts/extract_holy_grail.py

Output:
    quant-lab/reports/predecessor/extracted/
        ├── sheet_data/          — Per-sheet structured data (JSON)
        ├── price_data/          — Raw price CSVs (OHLCV)
        ├── ontology_stats/      — Stats mapped to ontology concepts
        ├── summary.json         — Master summary of all extractions
        └── extraction_log.md    — Human-readable extraction report
"""

import openpyxl
import json
import os
import csv
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict

# ── CONFIG ────────────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx"
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "reports", "predecessor", "extracted"
)
PRICE_DATA_DIR = os.path.join(OUTPUT_DIR, "price_data")
SHEET_DATA_DIR = os.path.join(OUTPUT_DIR, "sheet_data")
ONTOLOGY_DIR = os.path.join(OUTPUT_DIR, "ontology_stats")

# Sheets known to contain raw price data (OHLCV format)
PRICE_DATA_SHEETS = {
    "OILUSD_H4_202001020000_202505162000",
    "OILUSD_H1_202001020000_202505162300",
    "EURUSD_Daily_202001020000_202512180000",
    "EURUSD_Weekly_202001050000_202512140000",
    "EURUSD_H4_202001020000_202512181600",
    "EURUSD_H1_202001020000_202512181800",
    "DAILY DELIVERY NAVIGATION",
    "ETH_RANGE_EXPLORATION",
    "ETH_M15_DATA",
    "ETH_H1_DATA",
    "PHASE 4C - THRESHOLD TEST RESULTS",
}

# Sheets with tabular/analysis data (not raw price)
ANALYSIS_SHEETS = {
    "PHASE 2 VALIDATION RESULTS",
    "Delivery Stats",
    "Pattern Formations",
    "Pattern Failures & Rekeys",
    "ILM Zone Behaviors",
    "Session & Timing Metrics",
    "Fibonacci Sequences Catalog",
    "Validation Checklist",
    "Failure Pattern Database",
    "Low-Freq High-Accuracy Tracker",
    "Hit Rate Analysis Framework",
    "monday_fibonacci_calculations",
    "hit_rate_summary",
    "top5_claims_validation",
    "measurement_comparison",
    "quarterly_analysis",
    "session_data_full_week",
    "thursday_range_targets",
    "previous_day_targets",
    "EURUSD_Monday_Fibonacci",
    "EURUSD_Oil_Comparison",
    "EURUSD_132_PATTERNS",
    "EURUSD_TOLERANCE_ANALYSIS",
    "EURUSD_WEEKLY_REKEYS",
    "EURUSD_TEMPORAL_PATTERNS",
    "EURUSD_DAILY_REKEYS",
    "EURUSD_Asian_Fibonacci",
    "EURUSD_Asian_Hit_Rates",
    "EURUSD_Asian_Failures",
    "ASIAN→LONDON ALGO - PHASE 1",
    "ETH_Monday_Fibonacci",
    "ETH_Sunday_Asian",
    "ETH_Model_Comparison",
    "ETH_Analysis_Summary",
    "ETH_Fib_Analysis",
    "eth_friday_asian_ranges",
    "eth_fibonacci_hit_results",
    "eth_fibonacci_timing_analysis",
    "eth_session_probabilities",
    "eth_132_violations",
    "ETH_Data_Summary",
    "ETH_Model_Compilation",
    "Cross_Market_Comparison",
    "ETH_Discrepancy_Analysis",
}


def safe_str(val: Any) -> str:
    """Convert any value to a safe string."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return str(val).strip()


def safe_float(val: Any) -> Optional[float]:
    """Convert value to float, return None if not possible."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        # Handle strings like "85.4%" or "1,234.56"
        s = str(val).strip().replace(",", "").replace("%", "")
        return float(s)
    except (ValueError, TypeError):
        return None


def detect_table_start(ws, max_scan_rows: int = 5) -> Tuple[int, int]:
    """
    Detect where the actual data table starts in a sheet.
    Returns (header_row, data_start_row).
    """
    for row_idx in range(1, min(max_scan_rows + 1, ws.max_row + 1)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is not None:
            cell_str = str(cell_val).strip()
            # Check if first row looks like headers
            if row_idx == 1:
                return (1, 2)
            # Check for common header patterns
            if any(kw in cell_str.lower() for kw in ["date", "metric", "claim", "pattern", "session", "fib", "level", "hit", "bias", "range", "open", "high", "low", "close"]):
                return (row_idx, row_idx + 1)
    return (1, 2)


def extract_structured_table(ws, sheet_name: str) -> Dict:
    """
    Extract a structured table from a worksheet.
    Handles merged cells, multi-row headers, and irregular layouts.
    """
    header_row, data_start = detect_table_start(ws)
    
    # Extract headers
    headers = []
    max_col = min(ws.max_column or 1, 50)  # Cap at 50 columns
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        headers.append(safe_str(val) if val else f"col_{col_idx}")
    
    # Extract data rows
    rows = []
    max_row = min(ws.max_row or 0, 10000)  # Cap at 10k rows
    for row_idx in range(data_start, max_row + 1):
        row_data = {}
        has_data = False
        for col_idx in range(1, len(headers) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                has_data = True
                header = headers[col_idx - 1]
                # Try to convert to number
                fval = safe_float(val)
                row_data[header] = fval if fval is not None else safe_str(val)
        if has_data:
            rows.append(row_data)
    
    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "headers": headers,
        "row_count": len(rows),
        "data": rows,
    }


def extract_text_blocks(ws, sheet_name: str) -> Dict:
    """
    Extract text-heavy sheets as structured text blocks.
    Used for analysis sheets with embedded paragraphs and tables.
    """
    blocks = []
    current_block = {"type": "text", "content": [], "row": 0}
    
    max_row = min(ws.max_row or 0, 10000)
    for row_idx in range(1, max_row + 1):
        row_values = []
        max_col = min(ws.max_column or 1, 30)
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                row_values.append(safe_str(val))
        
        if row_values:
            line = " | ".join(row_values)
            # Detect section headers
            if any(kw in line.upper() for kw in ["SECTION", "PHASE", "PART", "SUMMARY", "ANALYSIS", "CONCLUSION", "RESULT"]):
                if current_block["content"]:
                    blocks.append(current_block)
                current_block = {"type": "header", "content": [line], "row": row_idx}
            else:
                current_block["content"].append(line)
                if current_block["type"] == "header":
                    current_block["type"] = "section"
    
    if current_block["content"]:
        blocks.append(current_block)
    
    return {
        "sheet_name": sheet_name,
        "block_count": len(blocks),
        "blocks": blocks,
    }


def extract_price_data(ws, sheet_name: str) -> Optional[str]:
    """
    Extract raw OHLCV price data to CSV.
    Returns the path to the created CSV file.
    """
    # Detect header row
    header_row = 1
    first_cell = ws.cell(row=1, column=1).value
    if first_cell and "DATE" in str(first_cell).upper():
        header_row = 1
    else:
        # Scan for header
        for r in range(1, min(5, ws.max_row + 1)):
            cell = ws.cell(row=r, column=1).value
            if cell and ("DATE" in str(cell).upper() or "OPEN" in str(cell).upper()):
                header_row = r
                break
    
    # Determine column mapping
    date_col = 1
    time_col = 2
    open_col = 3
    high_col = 4
    low_col = 5
    close_col = 6
    vol_col = 7
    spread_col = 9
    
    # Check if there's a time column
    has_time = False
    for c in range(1, min(10, ws.max_column + 1)):
        header_val = ws.cell(row=header_row, column=c).value
        if header_val and "TIME" in str(header_val).upper():
            has_time = True
            time_col = c
            break
    
    # Build CSV
    safe_name = re.sub(r'[^\w\-]', '_', sheet_name)[:80]
    csv_path = os.path.join(PRICE_DATA_DIR, f"{safe_name}.csv")
    
    rows_written = 0
    max_row = min(ws.max_row or 0, 100000)
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["date", "time", "open", "high", "low", "close", "tickvol", "spread"])
        
        for row_idx in range(header_row + 1, max_row + 1):
            date_val = ws.cell(row=row_idx, column=date_col).value
            if date_val is None:
                continue
            
            time_val = ws.cell(row=row_idx, column=time_col).value if has_time else ""
            open_val = safe_float(ws.cell(row=row_idx, column=open_col).value)
            high_val = safe_float(ws.cell(row=row_idx, column=high_col).value)
            low_val = safe_float(ws.cell(row=row_idx, column=low_col).value)
            close_val = safe_float(ws.cell(row=row_idx, column=close_col).value)
            tickvol = safe_float(ws.cell(row=row_idx, column=vol_col).value)
            spread = safe_float(ws.cell(row=row_idx, column=spread_col).value)
            
            if open_val is None or high_val is None or low_val is None or close_val is None:
                continue
            
            writer.writerow([
                safe_str(date_val),
                safe_str(time_val) if time_val else "",
                open_val, high_val, low_val, close_val,
                tickvol or 0, spread or 0
            ])
            rows_written += 1
    
    return csv_path if rows_written > 0 else None


def extract_key_stats_from_sheet(ws, sheet_name: str) -> Dict:
    """
    Extract key numerical stats from a sheet by scanning for patterns.
    Looks for: percentages, hit rates, counts, ratios, pips.
    """
    stats = {}
    max_row = min(ws.max_row or 0, 5000)
    max_col = min(ws.max_column or 50, 50)
    
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            
            val_str = safe_str(val)
            
            # Pattern: "XX.X%" or "XX%"
            pct_match = re.search(r'(\d+\.?\d*)%', val_str)
            if pct_match:
                pct_val = float(pct_match.group(1))
                # Get context from same row
                context_parts = []
                for c in range(1, min(col_idx, max_col + 1)):
                    ctx = ws.cell(row=row_idx, column=c).value
                    if ctx and c != col_idx:
                        ctx_str = safe_str(ctx)
                        if ctx_str and len(ctx_str) < 100:
                            context_parts.append(ctx_str)
                context = " | ".join(context_parts[-3:])  # Last 3 columns as context
                
                stat_key = f"pct_{row_idx}_{col_idx}"
                if context:
                    # Clean context for use as key
                    clean_ctx = re.sub(r'[^\w\s]', '', context)[:60].strip().replace(' ', '_')
                    stat_key = clean_ctx if clean_ctx else stat_key
                
                stats[stat_key] = {
                    "value": pct_val,
                    "type": "percentage",
                    "context": context,
                    "row": row_idx,
                    "col": col_idx,
                }
            
            # Pattern: "X,XXX" or "XXXX" (counts)
            count_match = re.search(r'^([\d,]+)$', val_str.replace(" ", ""))
            if count_match:
                count_val = int(count_match.group(1).replace(",", ""))
                if count_val > 10:  # Only meaningful counts
                    context_parts = []
                    for c in range(1, min(col_idx, max_col + 1)):
                        ctx = ws.cell(row=row_idx, column=c).value
                        if ctx and c != col_idx:
                            ctx_str = safe_str(ctx)
                            if ctx_str and len(ctx_str) < 100:
                                context_parts.append(ctx_str)
                    context = " | ".join(context_parts[-3:])
                    
                    stat_key = f"count_{row_idx}_{col_idx}"
                    clean_ctx = re.sub(r'[^\w\s]', '', context)[:60].strip().replace(' ', '_')
                    if clean_ctx:
                        stat_key = clean_ctx
                    
                    stats[stat_key] = {
                        "value": count_val,
                        "type": "count",
                        "context": context,
                        "row": row_idx,
                        "col": col_idx,
                    }
    
    return stats


def classify_sheet_type(sheet_name: str, ws) -> str:
    """Classify a sheet into one of: price_data, analysis_table, text_analysis, mixed."""
    if sheet_name in PRICE_DATA_SHEETS:
        return "price_data"
    
    # Check first few rows for patterns
    has_price_headers = False
    has_text_blocks = False
    has_table_structure = False
    
    for row_idx in range(1, min(5, ws.max_row + 1)):
        for col_idx in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                val_str = str(val).upper()
                if any(kw in val_str for kw in ["OPEN", "HIGH", "LOW", "CLOSE", "DATE", "TIME"]):
                    has_price_headers = True
                if len(str(val)) > 80:
                    has_text_blocks = True
    
    # Check for table structure (multiple rows with similar column counts)
    col_counts = []
    for row_idx in range(1, min(20, ws.max_row + 1)):
        count = sum(1 for c in range(1, min(20, ws.max_column + 1)) 
                   if ws.cell(row=row_idx, column=c).value is not None)
        col_counts.append(count)
    
    if len(col_counts) > 3:
        avg_cols = sum(col_counts) / len(col_counts)
        if avg_cols > 2 and max(col_counts) - min(col_counts) < 5:
            has_table_structure = True
    
    if has_price_headers:
        return "price_data"
    elif has_table_structure and not has_text_blocks:
        return "analysis_table"
    elif has_text_blocks:
        return "text_analysis"
    else:
        return "mixed"


def main():
    """Main extraction routine."""
    print("=" * 70)
    print("CEREBUS Holy Grail Excel Extractor")
    print("=" * 70)
    print(f"Source: {EXCEL_PATH}")
    print(f"Output: {OUTPUT_DIR}")
    print()
    
    # Create output directories
    os.makedirs(SHEET_DATA_DIR, exist_ok=True)
    os.makedirs(PRICE_DATA_DIR, exist_ok=True)
    os.makedirs(ONTOLOGY_DIR, exist_ok=True)
    
    # Load workbook
    print("Loading workbook (read-only mode)...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Total sheets: {len(wb.sheetnames)}")
    print()
    
    # Master summary
    summary = {
        "extraction_time": datetime.now().isoformat(),
        "source_file": EXCEL_PATH,
        "total_sheets": len(wb.sheetnames),
        "sheets": {},
        "price_data_files": [],
        "total_data_points": 0,
        "total_stats_extracted": 0,
    }
    
    # Process each sheet
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        sheet_type = classify_sheet_type(sheet_name, ws)
        
        print(f"[{i+1:3d}/{len(wb.sheetnames)}] {sheet_name[:60]:60s} type={sheet_type}", end="")
        
        sheet_info = {
            "index": i,
            "type": sheet_type,
            "rows": ws.max_row or 0,
            "cols": ws.max_column or 0,
        }
        
        try:
            if sheet_type == "price_data":
                # Extract to CSV
                csv_path = extract_price_data(ws, sheet_name)
                if csv_path:
                    rows = sum(1 for _ in open(csv_path, encoding='utf-8')) - 1
                    sheet_info["csv_file"] = os.path.basename(csv_path)
                    sheet_info["price_rows"] = rows
                    summary["price_data_files"].append({
                        "sheet": sheet_name,
                        "file": os.path.basename(csv_path),
                        "rows": rows,
                    })
                    summary["total_data_points"] += rows
                    print(f" -> {rows:,} price rows")
                else:
                    print(" -> NO DATA")
                    
            elif sheet_type == "analysis_table":
                # Extract structured table
                table_data = extract_structured_table(ws, sheet_name)
                json_path = os.path.join(SHEET_DATA_DIR, f"sheet_{i:03d}.json")
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(table_data, f, indent=2, default=str)
                
                sheet_info["json_file"] = f"sheet_{i:03d}.json"
                sheet_info["data_rows"] = table_data["row_count"]
                sheet_info["headers"] = table_data["headers"][:15]
                summary["total_data_points"] += table_data["row_count"]
                
                # Also extract key stats
                stats = extract_key_stats_from_sheet(ws, sheet_name)
                if stats:
                    stats_path = os.path.join(ONTOLOGY_DIR, f"stats_{i:03d}.json")
                    with open(stats_path, 'w', encoding='utf-8') as f:
                        json.dump(stats, f, indent=2, default=str)
                    sheet_info["stats_count"] = len(stats)
                    summary["total_stats_extracted"] += len(stats)
                
                print(f" -> {table_data['row_count']:,} rows, {len(stats)} stats")
                
            elif sheet_type == "text_analysis":
                # Extract text blocks
                text_data = extract_text_blocks(ws, sheet_name)
                json_path = os.path.join(SHEET_DATA_DIR, f"sheet_{i:03d}.json")
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(text_data, f, indent=2, default=str)
                
                sheet_info["json_file"] = f"sheet_{i:03d}.json"
                sheet_info["blocks"] = text_data["block_count"]
                
                # Also extract key stats from text
                stats = extract_key_stats_from_sheet(ws, sheet_name)
                if stats:
                    stats_path = os.path.join(ONTOLOGY_DIR, f"stats_{i:03d}.json")
                    with open(stats_path, 'w', encoding='utf-8') as f:
                        json.dump(stats, f, indent=2, default=str)
                    sheet_info["stats_count"] = len(stats)
                    summary["total_stats_extracted"] += len(stats)
                
                print(f" -> {text_data['block_count']} blocks, {len(stats)} stats")
                
            else:  # mixed
                # Try both approaches
                table_data = extract_structured_table(ws, sheet_name)
                stats = extract_key_stats_from_sheet(ws, sheet_name)
                
                json_path = os.path.join(SHEET_DATA_DIR, f"sheet_{i:03d}.json")
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(table_data, f, indent=2, default=str)
                
                sheet_info["json_file"] = f"sheet_{i:03d}.json"
                sheet_info["data_rows"] = table_data["row_count"]
                
                if stats:
                    stats_path = os.path.join(ONTOLOGY_DIR, f"stats_{i:03d}.json")
                    with open(stats_path, 'w', encoding='utf-8') as f:
                        json.dump(stats, f, indent=2, default=str)
                    sheet_info["stats_count"] = len(stats)
                    summary["total_stats_extracted"] += len(stats)
                
                summary["total_data_points"] += table_data["row_count"]
                print(f" -> {table_data['row_count']:,} rows, {len(stats)} stats")
        
        except Exception as e:
            sheet_info["error"] = str(e)
            print(f" -> ERROR: {e}")
        
        summary["sheets"][sheet_name] = sheet_info
    
    wb.close()
    
    # Save master summary
    summary_path = os.path.join(OUTPUT_DIR, "summary.json")
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, default=str)
    
    # Generate extraction log
    log_path = os.path.join(OUTPUT_DIR, "extraction_log.md")
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("# CEREBUS Holy Grail — Extraction Report\n\n")
        f.write(f"**Extracted:** {summary['extraction_time']}\n")
        f.write(f"**Source:** `{EXCEL_PATH}`\n")
        f.write(f"**Total Sheets:** {summary['total_sheets']}\n")
        f.write(f"**Total Data Points:** {summary['total_data_points']:,}\n")
        f.write(f"**Total Stats Extracted:** {summary['total_stats_extracted']:,}\n\n")
        
        f.write("## Price Data Files\n\n")
        f.write("| Sheet | File | Rows |\n")
        f.write("|-------|------|------|\n")
        for pf in summary["price_data_files"]:
            f.write(f"| {pf['sheet'][:50]} | {pf['file']} | {pf['rows']:,} |\n")
        
        f.write("\n## Sheet Summary\n\n")
        f.write("| # | Sheet | Type | Rows | Cols | Data Points | Stats |\n")
        f.write("|---|-------|------|------|------|-------------|-------|\n")
        for name, info in summary["sheets"].items():
            rows = info.get("data_rows", info.get("price_rows", info.get("blocks", "-")))
            stats = info.get("stats_count", "-")
            error = info.get("error", "")
            f.write(f"| {info['index']+1} | {name[:45]} | {info['type']} | {rows} | {info['cols']} | {info.get('data_rows', '-')} | {stats} {error} |\n")
    
    print()
    print("=" * 70)
    print(f"EXTRACTION COMPLETE")
    print(f"  Total data points: {summary['total_data_points']:,}")
    print(f"  Total stats extracted: {summary['total_stats_extracted']:,}")
    print(f"  Price data files: {len(summary['price_data_files'])}")
    print(f"  Summary: {summary_path}")
    print(f"  Log: {log_path}")
    print("=" * 70)


if __name__ == "__main__":
    main()

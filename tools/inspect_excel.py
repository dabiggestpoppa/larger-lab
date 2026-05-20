import openpyxl
from openpyxl.utils import get_column_letter
import json
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r'C:\Users\wifik\Downloads\cerebus 3 market hoily grail.xlsx', read_only=True, data_only=True)

print(f"Sheet count: {len(wb.sheetnames)}")
print(f"Sheets: {wb.sheetnames}\n")

# First pass: just get sheet names, row counts, and column headers
summary = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = []
    row_count = 0
    for i, row in enumerate(ws.iter_rows(max_row=min(5, ws.max_row), values_only=True)):
        row_count += 1
        if i == 0:
            headers = [str(v) if v is not None else "" for v in row[:20]]
    
    summary[sheet_name] = {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "headers": headers
    }

wb.close()

# Print summary
for name, info in summary.items():
    print(f"\n--- {name} ---")
    print(f"  Rows: {info['max_row']}, Cols: {info['max_col']}")
    print(f"  Headers: {info['headers'][:15]}")

# Save full summary to JSON
with open(r'C:\Users\wifik\Desktop\projects\larger-lab\quant-lab\research\excel_structure_summary.json', 'w', encoding='utf-8') as f:
    json.dump(summary, f, indent=2, ensure_ascii=False, default=str)

print(f"\n\nFull summary saved to quant-lab/research/excel_structure_summary.json")

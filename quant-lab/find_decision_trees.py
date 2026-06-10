import openpyxl, sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx", read_only=True)

keywords = [
    "decision tree", "playbook", "phase 4", "phase 4a", "phase 4b", "phase 4c",
    "entry rule", "exit rule", "trigger", "condition", "if price", "then",
    "buy", "sell", "long", "short", "signal", "alert", "setup", "invalidation"
]

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_str = " ".join([str(v).lower() for v in row if v is not None])
        if any(kw in row_str for kw in keywords):
            non_empty = [str(v) for v in row if v is not None]
            if non_empty and len(non_empty) > 1:
                print(f"[{sheet_name}] row {i}: {non_empty[:12]}")

wb.close()

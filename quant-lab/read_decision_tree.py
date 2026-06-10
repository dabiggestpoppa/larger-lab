import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx", read_only=True)

# Read ALL sheets that might contain decision trees or playbooks
target_sheets = [
    "DECISION TREE - WEEKLY CLOSE",
    "PHASE 4 - MONTHLY RANGE RECONNAISSANCE",
    "PHASE 4A - MONTHLY DATASET",
    "PHASE 4B - RANGE WINDOW TESTING",
    "PHASE 4C - GROUP B ANALYSIS",
    "PHASE 4 - TEMPORAL DELIVERY MAPPING",
    "PHASE 5 - WILM ILM VELOCITY ANALYSIS",
    "PHASE 6 - SESSION PROFILE SYNTHESIS",
    "Validation Checklist",
    "Top 10 Claims - Testing Framework",
    "Hit Rate Analysis Framework",
]

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"  SHEET NOT FOUND: {sheet_name}")
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"=== {sheet_name} ===")
    print(f"{'='*70}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        non_empty = [str(v) for v in row if v is not None]
        if non_empty:
            print(f"  row {i}: {non_empty[:15]}")

wb.close()

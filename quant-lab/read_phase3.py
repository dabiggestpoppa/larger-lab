import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx", read_only=True)

# PHASE 3 - Comprehensive Analysis
ws = wb["PHASE 3 - Comprehensive Analysis"]
print("=== PHASE 3 - Comprehensive Analysis (rows 85-130) ===")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 85:
        continue
    if i > 130:
        break
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        print(f"  row {i}: {non_empty[:12]}")

print()
print("=== PHASE 3B - Temporal Delivery System (rows 45-80) ===")
ws2 = wb["PHASE 3B - Temporal Delivery System"]
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i < 45:
        continue
    if i > 80:
        break
    non_empty = [str(v) for v in row if v is not None]
    if non_empty:
        print(f"  row {i}: {non_empty[:12]}")

wb.close()

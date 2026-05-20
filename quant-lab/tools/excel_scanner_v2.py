"""
CEREBUS Excel Scanner v2 - Handles read_only mode limitations
"""
import openpyxl
import json
import sys

FILE_PATH = r'C:\Users\wifik\Downloads\cerebus 3 market hoily grail.xlsx'
OUTPUT_PATH = r'C:\Users\wifik\Desktop\projects\larger-lab\quant-lab\research\excel_scan_results.json'

def scan_sheet(ws, max_rows=20):
    """Scan a single sheet and return its structure."""
    headers = []
    sample_rows = []
    row_count = 0
    max_col = 0
    max_row_count = 0
    
    for row in ws.iter_rows(max_row=max_rows, values_only=False):
        row_count += 1
        max_row_count = row_count
        vals = []
        for cell in row:
            if cell.value is not None:
                vals.append((cell.coordinate, cell.value))
                max_col = max(max_col, cell.column)
        if row_count == 1:
            headers = vals
        else:
            if vals:
                sample_rows.append(vals)
    
    return {
        'name': ws.title,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'scanned_rows': max_row_count,
        'headers': [(c, v) for c, v in headers],
        'sample_rows': [[(c, v) for c, v in row] for row in sample_rows],
    }

def main():
    print("Loading workbook (read_only=True, data_only=True)...")
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    print(f"Loaded {len(wb.sheetnames)} sheets")
    
    # Focus on sheets 1-42
    target_sheets = wb.sheetnames[:42]
    
    results = {}
    for i, name in enumerate(target_sheets):
        idx = i + 1
        print(f"  Scanning [{idx}] {name}...")
        try:
            ws = wb[name]
            results[f"{idx:02d}_{name}"] = scan_sheet(ws)
        except Exception as e:
            results[f"{idx:02d}_{name}"] = {'error': str(e)}
    
    wb.close()
    
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, default=str, ensure_ascii=False)
    
    print(f"\nDone. Results saved to {OUTPUT_PATH}")
    print(f"Total sheets scanned: {len(results)}")

if __name__ == '__main__':
    main()

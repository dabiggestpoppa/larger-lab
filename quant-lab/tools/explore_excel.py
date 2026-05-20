import openpyxl
import json
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\wifik\Downloads\cerebus 3 market hoily grail.xlsx', read_only=True, data_only=True)

output_dir = r'C:\Users\wifik\Desktop\projects\larger-lab\quant-lab\research'
os.makedirs(output_dir, exist_ok=True)

def explore_sheet(wb, sheet_name, max_rows=30):
    """Extract headers and first few rows from a sheet"""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"sheet": sheet_name, "empty": True}
    
    # Find header row (first non-empty row)
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 3:
            header_row = row
            header_idx = i
            break
    
    if header_row is None:
        header_row = rows[0]
        header_idx = 0
    
    # Get data rows after header
    data_start = header_idx + 1
    data_rows = []
    count = 0
    for row in rows[data_start:]:
        if any(c is not None for c in row):
            data_rows.append(row)
            count += 1
            if count >= max_rows:
                break
    
    # Count total data rows
    total_data = sum(1 for row in rows[data_start:] if any(c is not None for c in row))
    
    return {
        "sheet": sheet_name,
        "headers": [str(h) if h is not None else "" for h in header_row],
        "data_sample": [[str(v) if v is not None else "" for v in row] for row in data_rows[:10]],
        "total_rows": len(rows),
        "data_rows": total_data,
        "header_row_index": header_idx
    }

# Key sheets to explore in detail
key_sheets = [
    "Delivery Stats",
    "Pattern Formations",
    "Pattern Failures & Rekeys",
    "ILM Zone Behaviors",
    "Session & Timing Metrics",
    "Fibonacci Sequences Catalog",
    "Validation Checklist",
    "Failure Pattern Database",
    "Hit Rate Analysis Framework",
    "monday_fibonacci_calculations",
    "hit_rate_summary",
    "session_data_full_week",
    "PHASE 2 - OVERLAP ANALYSIS",
    "TOLERANCE_COMPARISON_0.15_0.25_0.50",
    "MONDAY-ASIAN SESSION CROSS-REF",
    "ASIAN SESSION FAILURE ANALYSIS",
    "PHASE 3 - SESSION CORRELATION MATRIX",
    "PHASE 4 - TEMPORAL DELIVERY MAPPING",
    "PHASE 5 - WILM ILM VELOCITY ANALYSIS",
    "PHASE 6 - SESSION PROFILE SYNTHESIS",
    "PHASE 7 - MODEL SYNTHESIS & INTEGRATION",
    "REKEY HYPOTHESIS TEST RESULTS",
    "Cross_Market_Comparison",
    "EURUSD_TOLERANCE_ANALYSIS",
    "EURUSD_WEEKLY_REKEYS",
    "EURUSD_Asian_Hit_Rates",
    "EURUSD_DAILY_REKEYS",
    "ETH_Session_Probabilities",
    "ETH_Fib_Analysis",
    "ETH_Range_Exploration",
    "previous_day_targets",
    "thursday_range_targets",
    "quarterly_analysis",
    "measurement_comparison",
    "top5_claims_validation",
    "Low-Freq High-Accuracy Tracker",
    "Top 10 Claims - Testing Framework",
    "PHASE 2 VALIDATION RESULTS",
]

all_data = {}
for sheet_name in key_sheets:
    if sheet_name in wb.sheetnames:
        print(f"\n=== {sheet_name} ===")
        info = explore_sheet(wb, sheet_name)
        all_data[sheet_name] = info
        print(f"  Headers ({len(info['headers'])}): {info['headers'][:15]}")
        print(f"  Data rows: {info['data_rows']}")
        if info['data_sample']:
            print(f"  Sample row: {info['data_sample'][0][:10]}")
    else:
        print(f"\n=== {sheet_name} NOT FOUND ===")

# Save all data as JSON
with open(os.path.join(output_dir, 'excel_structure.json'), 'w', encoding='utf-8') as f:
    json.dump(all_data, f, indent=2, ensure_ascii=False, default=str)

wb.close()
print(f"\n\nDone. Explored {len(all_data)} sheets. Data saved to excel_structure.json")

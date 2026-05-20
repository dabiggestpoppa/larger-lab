"""
CEREBUS Stat Tracking System Extractor
Extracts MAD's stat tracking methodology from the Excel file.
"""
import openpyxl
import json
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail.xlsx"
OUTPUT_DIR = r"C:\Users\wifik\Desktop\projects\larger-lab\quant-lab\data"

def safe_str(val):
    if val is None:
        return ""
    return str(val).replace('\n', ' ').strip()

def read_sheet(ws, max_rows=200):
    """Read a sheet into a list of lists, cleanly."""
    rows = []
    count = 0
    for row in ws.iter_rows(values_only=True, max_row=max_rows):
        rows.append([safe_str(c) for c in row])
        count += 1
    return rows

def extract_sheet_data(wb, sheet_name, max_rows=200):
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    return {
        "name": sheet_name,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "data": read_sheet(ws, max_rows)
    }

def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Loaded {len(wb.sheetnames)} sheets")
    
    # Key sheets for stat tracking methodology
    key_sheets = [
        "hit_rate_summary",
        "Delivery Stats", 
        "Pattern Formations",
        "Pattern Failures & Rekeys",
        "Session & Timing Metrics",
        "Fibonacci Sequences Catalog",
        "failure_pattern_database",
        "REKEY HYPOTHESIS TEST RESULTS",
        "ASIAN SESSION FAILURE ANALYSIS",
        "monday_fibonacci_calculations",
        "session_data_full_week",
        "thursday_range_targets",
        "previous_day_targets",
        "TOLERANCE_COMPARISON_0.15_0.25_0.50",
        "PHASE 2 - OVERLAP ANALYSIS",
        "PHASE 3 - SESSION CORRELATION MATRIX",
        "PHASE 6 - SESSION PROFILE SYNTHESIS",
        "PHASE 2 VALIDATION RESULTS",
        "ILM Zone Behaviors",
        "Low-Freq High-Accuracy Tracker",
        "Top 10 Claims - Testing Framework",
        "Hit Rate Analysis Framework",
        "top5_claims_validation",
        "measurement_comparison",
        "quarterly_analysis",
    ]
    
    all_data = {}
    for sname in key_sheets:
        data = extract_sheet_data(wb, sname, max_rows=100)
        if data:
            all_data[sname] = data
            print(f"  [{sname}] {data['rows']}x{data['cols']}")
        else:
            print(f"  [{sname}] NOT FOUND")
    
    # Save as JSON for further processing
    output_path = os.path.join(OUTPUT_DIR, "cerebus_key_sheets.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nSaved to {output_path}")
    
    wb.close()
    print("Done.")

if __name__ == "__main__":
    main()

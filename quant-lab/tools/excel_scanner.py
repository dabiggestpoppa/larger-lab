"""
CEREBUS Excel Scanner - Extracts structure from all sheets (1-42 focus)
Outputs JSON for further processing.
"""
import openpyxl
import json
import sys
import os

FILE_PATH = r'C:\Users\wifik\Downloads\cerebus 3 market hoily grail.xlsx'
OUTPUT_PATH = r'C:\Users\wifik\Desktop\projects\larger-lab\quant-lab\research\excel_scan_results.json'

def scan_sheet(ws, max_rows=15):
    """Scan a single sheet and return its structure."""
    result = {
        'name': ws.title,
        'dimensions': ws.dimensions,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'headers': [],
        'sample_rows': [],
        'merged_cells': [str(m) for m in ws.merged_cells.ranges[:10]],
    }
    
    row_count = 0
    for row in ws.iter_rows(max_row=min(max_rows, ws.max_row), values_only=False):
        row_count += 1
        vals = [(cell.coordinate, cell.value) for cell in row if cell.value is not None]
        if row_count == 1:
            result['headers'] = vals
        else:
            if vals:
                result['sample_rows'].append(vals)
    
    return result

def main():
    print(f"Loading workbook (read_only=True, data_only=True)...")
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    print(f"Loaded {len(wb.sheetnames)} sheets")
    
    # Focus on sheets 1-42
    target_sheets = wb.sheetnames[:42]
    
    results = {}
    for i, name in enumerate(target_sheets):
        idx = i + 1
        print(f"  Scanning [{idx}] {name}...")
        try:
            ws = wb[name]
            results[f"{idx:02d}_{name}"] = scan_sheet(ws)
        except Exception as e:
            results[f"{idx:02d}_{name}"] = {'error': str(e)}
    
    wb.close()
    
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, default=str, ensure_ascii=False)
    
    print(f"\nDone. Results saved to {OUTPUT_PATH}")
    print(f"Total sheets scanned: {len(results)}")

if __name__ == '__main__':
    main()

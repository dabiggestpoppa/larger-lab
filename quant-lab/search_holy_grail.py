import openpyxl, json, re

wb = openpyxl.load_workbook(r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx", read_only=True)

keywords = ["alpha", "beta", "gamma", "block", "72", "50-25", "sequence", "intraday", "intraweek", "weekly fib", "fibonacci sequence"]

results = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheetname] if hasattr(wb, 'sheetname') else wb[sheet_name]
    sheet_results = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_str = " ".join([str(v).lower() for v in row if v is not None])
        if any(kw in row_str for kw in keywords):
            non_empty = [str(v) for v in row if v is not None]
            if non_empty and len(non_empty) > 1:
                sheet_results.append({"row": i, "data": non_empty[:15]})
    if sheet_results:
        results[sheet_name] = sheet_results

# Print results
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

for sheet, rows in results.items():
    print(f"\n=== {sheet} ===")
    for r in rows[:10]:
        print(f"  row {r['row']}: {r['data']}")

wb.close()

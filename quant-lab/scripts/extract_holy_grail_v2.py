"""
CEREBUS Holy Grail Excel Extractor v2
======================================
Extracts EVERY stat from EVERY sheet in the predecessor Excel file.
Handles mixed formats: structured tables, embedded analysis, raw price data.
Outputs structured JSON for ontology ingestion.

Key improvement: Better sheet classification based on actual header patterns.

Usage:
    python quant-lab/scripts/extract_holy_grail_v2.py
"""

import openpyxl
import json
import os
import csv
import re
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

# ── CONFIG ────────────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx"
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "reports", "predecessor", "extracted"
)
PRICE_DATA_DIR = os.path.join(OUTPUT_DIR, "price_data")
SHEET_DATA_DIR = os.path.join(OUTPUT_DIR, "sheet_data")
ONTOLOGY_DIR = os.path.join(OUTPUT_DIR, "ontology_stats")


def safe_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return str(val).strip()


def safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace(",", "").replace("%", "")
        return float(s)
    except (ValueError, TypeError):
        return None


def is_price_data_sheet(sheet_name: str, ws) -> bool:
    """
    Determine if a sheet contains raw OHLCV price data.
    Checks: sheet name patterns + first row header patterns.
    """
    # Check sheet name patterns
    price_name_patterns = [
        r'_(H1|H4|M15|M5|D1|W1|MN)_',  # Timeframe patterns
        r'_RAW_DATA$',
        r'_Daily_',
        r'_Weekly_',
        r'ETH_(M15|H1|DATA)',
        r'OILUSD_(H1|H4)',
        r'DAILY DELIVERY NAVIGATION',
        r'ETH_RANGE_EXPLORATION',
        r'PHASE 4C - THRESHOLD TEST RESULTS',
    ]
    for pattern in price_name_patterns:
        if re.search(pattern, sheet_name, re.IGNORECASE):
            return True
    
    # Check first row for OHLCV header patterns
    first_row = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        first_row = [str(v).upper() if v else "" for v in row]
    
    first_cell = first_row[0] if first_row else ""
    
    # Price data sheets have specific first-cell patterns
    if '<DATE>' in first_cell or '<OPEN>' in first_cell:
        return True
    if 'DATE' in first_cell and ('TIME' in first_cell or 'OPEN' in first_cell):
        return True
    if first_cell.startswith('DATE') and len(first_row) > 5:
        # Check if columns 2-6 look like OHLCV
        if any('OPEN' in c for c in first_row[1:5]):
            return True
        if any('HIGH' in c for c in first_row[1:5]):
            return True
    
    return False


def extract_price_data_to_csv(ws, sheet_name: str) -> Optional[str]:
    """Extract raw OHLCV price data to CSV file."""
    # Detect header row and column positions
    header_row_idx = 1
    
    # Read first row to understand structure
    first_row = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        first_row = list(row)
        break
    
    if not first_row:
        return None
    
    first_cell = str(first_row[0]) if first_row[0] else ""
    
    # Determine column positions
    date_col = 0  # 0-indexed
    time_col = -1
    open_col = 2
    high_col = 3
    low_col = 4
    close_col = 5
    vol_col = 6
    spread_col = 8
    
    # Check if first cell contains tab-separated headers (MT5 export format)
    if '\t' in first_cell or '_x0009_' in first_cell:
        # MT5 export format: <DATE>\t<TIME>\t<OPEN>\t<HIGH>\t<LOW>\t<CLOSE>\t<TICKVOL>\t<VOL>\t<SPREAD>
        # These are in a single cell separated by tabs
        time_col = 1
    else:
        # Check individual cells for column headers
        for i, val in enumerate(first_row[:15]):
            val_str = str(val).upper() if val else ""
            if 'DATE' in val_str:
                date_col = i
            elif 'TIME' in val_str:
                time_col = i
            elif 'OPEN' in val_str and 'HIGH' not in val_str:
                open_col = i
            elif 'HIGH' in val_str:
                high_col = i
            elif 'LOW' in val_str:
                low_col = i
            elif 'CLOSE' in val_str:
                close_col = i
            elif 'TICKVOL' in val_str or 'VOL' in val_str:
                vol_col = i
            elif 'SPREAD' in val_str:
                spread_col = i
    
    # Build CSV
    safe_name = re.sub(r'[^\w\-]', '_', sheet_name)[:80]
    csv_path = os.path.join(PRICE_DATA_DIR, f"{safe_name}.csv")
    
    rows_written = 0
    max_row = min(ws.max_row or 0, 100000)
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["date", "time", "open", "high", "low", "close", "tickvol", "spread"])
        
        for row_idx in range(2, max_row + 1):
            row_data = []
            max_col = min(ws.max_column or 0, 20)
            for col_idx in range(1, max_col + 1):
                row_data.append(ws.cell(row=row_idx, column=col_idx).value)
            
            if not row_data or row_data[0] is None:
                continue
            
            # Handle MT5 tab-export format (all data in first cell)
            if time_col == 1 and '\t' in str(row_data[0]):
                parts = str(row_data[0]).split('\t')
                if len(parts) >= 6:
                    date_val = parts[0]
                    time_val = parts[1] if len(parts) > 1 else ""
                    try:
                        open_val = float(parts[2])
                        high_val = float(parts[3])
                        low_val = float(parts[4])
                        close_val = float(parts[5])
                        tickvol = float(parts[6]) if len(parts) > 6 else 0
                        spread = float(parts[8]) if len(parts) > 8 else 0
                    except (ValueError, IndexError):
                        continue
                    
                    writer.writerow([date_val, time_val, open_val, high_val, low_val, close_val, tickvol, spread])
                    rows_written += 1
            else:
                # Normal column-based format
                date_val = str(row_data[date_col]) if date_col < len(row_data) else ""
                time_val = str(row_data[time_col]) if time_col >= 0 and time_col < len(row_data) else ""
                open_val = safe_float(row_data[open_col]) if open_col < len(row_data) else None
                high_val = safe_float(row_data[high_col]) if high_col < len(row_data) else None
                low_val = safe_float(row_data[low_col]) if low_col < len(row_data) else None
                close_val = safe_float(row_data[close_col]) if close_col < len(row_data) else None
                tickvol = safe_float(row_data[vol_col]) if vol_col < len(row_data) else 0
                spread = safe_float(row_data[spread_col]) if spread_col < len(row_data) else 0
                
                if open_val is None or high_val is None or low_val is None or close_val is None:
                    continue
                
                writer.writerow([date_val, time_val, open_val, high_val, low_val, close_val, tickvol or 0, spread or 0])
                rows_written += 1
    
    return csv_path if rows_written > 0 else None


def extract_structured_table(ws, sheet_name: str) -> Dict:
    """Extract a structured table from a worksheet."""
    # Detect header row
    header_row = 1
    for row_idx in range(1, min(5, ws.max_row + 1)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is not None:
            header_row = row_idx
            break
    
    # Extract headers
    headers = []
    max_col = min(ws.max_column or 1, 50)
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        headers.append(safe_str(val) if val else f"col_{col_idx}")
    
    # Extract data rows
    rows = []
    max_row = min(ws.max_row or 0, 10000)
    for row_idx in range(header_row + 1, max_row + 1):
        row_data = {}
        has_data = False
        for col_idx in range(1, len(headers) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                has_data = True
                header = headers[col_idx - 1]
                fval = safe_float(val)
                row_data[header] = fval if fval is not None else safe_str(val)
        if has_data:
            rows.append(row_data)
    
    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "headers": headers,
        "row_count": len(rows),
        "data": rows,
    }


def extract_key_stats(ws, sheet_name: str) -> Dict:
    """Extract key numerical stats from a sheet."""
    stats = {}
    max_row = min(ws.max_row or 0, 5000)
    max_col = min(ws.max_column or 50, 50)
    
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            
            val_str = safe_str(val)
            
            # Pattern: percentage values
            pct_match = re.search(r'(\d+\.?\d*)%', val_str)
            if pct_match:
                pct_val = float(pct_match.group(1))
                # Get context from same row (preceding columns)
                context_parts = []
                for c in range(1, col_idx):
                    ctx = ws.cell(row=row_idx, column=c).value
                    if ctx:
                        ctx_str = safe_str(ctx)
                        if ctx_str and len(ctx_str) < 100:
                            context_parts.append(ctx_str)
                context = " | ".join(context_parts[-3:])
                
                stat_key = f"pct_r{row_idx}_c{col_idx}"
                clean_ctx = re.sub(r'[^\w\s]', '', context)[:60].strip().replace(' ', '_')
                if clean_ctx:
                    stat_key = clean_ctx
                
                stats[stat_key] = {
                    "value": pct_val,
                    "type": "percentage",
                    "context": context,
                    "row": row_idx,
                    "col": col_idx,
                }
            
            # Pattern: decimal values that look like rates/ratios (0.XXX)
            if isinstance(val, float) and 0 < val < 1:
                context_parts = []
                for c in range(1, col_idx):
                    ctx = ws.cell(row=row_idx, column=c).value
                    if ctx:
                        ctx_str = safe_str(ctx)
                        if ctx_str and len(ctx_str) < 100:
                            context_parts.append(ctx_str)
                context = " | ".join(context_parts[-3:])
                
                stat_key = f"ratio_r{row_idx}_c{col_idx}"
                clean_ctx = re.sub(r'[^\w\s]', '', context)[:60].strip().replace(' ', '_')
                if clean_ctx:
                    stat_key = clean_ctx
                
                stats[stat_key] = {
                    "value": val,
                    "type": "ratio",
                    "context": context,
                    "row": row_idx,
                    "col": col_idx,
                }
    
    return stats


def main():
    print("=" * 70)
    print("CEREBUS Holy Grail Excel Extractor v2")
    print("=" * 70)
    print(f"Source: {EXCEL_PATH}")
    print(f"Output: {OUTPUT_DIR}")
    print()
    
    os.makedirs(SHEET_DATA_DIR, exist_ok=True)
    os.makedirs(PRICE_DATA_DIR, exist_ok=True)
    os.makedirs(ONTOLOGY_DIR, exist_ok=True)
    
    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Total sheets: {len(wb.sheetnames)}")
    print()
    
    summary = {
        "extraction_time": datetime.now().isoformat(),
        "source_file": EXCEL_PATH,
        "total_sheets": len(wb.sheetnames),
        "sheets": {},
        "price_data_files": [],
        "total_data_points": 0,
        "total_stats_extracted": 0,
    }
    
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        is_price = is_price_data_sheet(sheet_name, ws)
        sheet_type = "price_data" if is_price else "analysis"
        
        print(f"[{i+1:3d}/{len(wb.sheetnames)}] {sheet_name[:55]:55s} [{sheet_type}]", end="", flush=True)
        
        sheet_info = {
            "index": i,
            "type": sheet_type,
            "rows": ws.max_row or 0,
            "cols": ws.max_column or 0,
        }
        
        try:
            if is_price:
                csv_path = extract_price_data_to_csv(ws, sheet_name)
                if csv_path:
                    rows = sum(1 for _ in open(csv_path, encoding='utf-8')) - 1
                    sheet_info["csv_file"] = os.path.basename(csv_path)
                    sheet_info["price_rows"] = rows
                    summary["price_data_files"].append({
                        "sheet": sheet_name,
                        "file": os.path.basename(csv_path),
                        "rows": rows,
                    })
                    summary["total_data_points"] += rows
                    print(f" -> {rows:,} price rows -> {os.path.basename(csv_path)}")
                else:
                    # Fall back to table extraction
                    print(" -> No price data found, extracting as table...")
                    table_data = extract_structured_table(ws, sheet_name)
                    json_path = os.path.join(SHEET_DATA_DIR, f"sheet_{i:03d}.json")
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(table_data, f, indent=2, default=str)
                    sheet_info["json_file"] = f"sheet_{i:03d}.json"
                    sheet_info["data_rows"] = table_data["row_count"]
                    summary["total_data_points"] += table_data["row_count"]
                    print(f"   -> {table_data['row_count']:,} table rows")
            else:
                # Extract structured table
                table_data = extract_structured_table(ws, sheet_name)
                json_path = os.path.join(SHEET_DATA_DIR, f"sheet_{i:03d}.json")
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(table_data, f, indent=2, default=str)
                
                sheet_info["json_file"] = f"sheet_{i:03d}.json"
                sheet_info["data_rows"] = table_data["row_count"]
                sheet_info["headers"] = table_data["headers"][:15]
                summary["total_data_points"] += table_data["row_count"]
                
                # Extract key stats
                stats = extract_key_stats(ws, sheet_name)
                if stats:
                    stats_path = os.path.join(ONTOLOGY_DIR, f"stats_{i:03d}.json")
                    with open(stats_path, 'w', encoding='utf-8') as f:
                        json.dump(stats, f, indent=2, default=str)
                    sheet_info["stats_count"] = len(stats)
                    summary["total_stats_extracted"] += len(stats)
                
                print(f" -> {table_data['row_count']:,} rows, {len(stats)} stats")
        
        except Exception as e:
            sheet_info["error"] = str(e)
            print(f" -> ERROR: {e}")
        
        summary["sheets"][sheet_name] = sheet_info
    
    wb.close()
    
    # Save master summary
    summary_path = os.path.join(OUTPUT_DIR, "summary.json")
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, default=str)
    
    # Generate extraction log
    log_path = os.path.join(OUTPUT_DIR, "extraction_log.md")
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("# CEREBUS Holy Grail — Extraction Report\n\n")
        f.write(f"**Extracted:** {summary['extraction_time']}\n")
        f.write(f"**Source:** `{EXCEL_PATH}`\n")
        f.write(f"**Total Sheets:** {summary['total_sheets']}\n")
        f.write(f"**Total Data Points:** {summary['total_data_points']:,}\n")
        f.write(f"**Total Stats Extracted:** {summary['total_stats_extracted']:,}\n\n")
        
        f.write("## Price Data Files\n\n")
        f.write("| Sheet | File | Rows |\n")
        f.write("|-------|------|------|\n")
        for pf in summary["price_data_files"]:
            f.write(f"| {pf['sheet'][:50]} | {pf['file']} | {pf['rows']:,} |\n")
        
        f.write("\n## Sheet Summary\n\n")
        f.write("| # | Sheet | Type | Rows | Cols | Data Points | Stats | Error |\n")
        f.write("|---|-------|------|------|------|-------------|-------|-------|\n")
        for name, info in summary["sheets"].items():
            rows = info.get("data_rows", info.get("price_rows", "-"))
            stats_c = info.get("stats_count", "-")
            error = info.get("error", "")
            f.write(f"| {info['index']+1} | {name[:40]:40s} | {info['type']:10s} | {rows} | {info['cols']} | {info.get('data_rows', '-')} | {stats_c} {error} |\n")
    
    print()
    print("=" * 70)
    print(f"EXTRACTION COMPLETE")
    print(f"  Total data points: {summary['total_data_points']:,}")
    print(f"  Total stats extracted: {summary['total_stats_extracted']:,}")
    print(f"  Price data files: {len(summary['price_data_files'])}")
    print(f"  Summary: {summary_path}")
    print("=" * 70)


if __name__ == "__main__":
    main()

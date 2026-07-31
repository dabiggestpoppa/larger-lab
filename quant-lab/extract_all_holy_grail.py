"""
Extract ALL data points from the Holy Grail Excel for ML model training.
Outputs a single JSON file with every stat, sequence, and data point.
"""
import openpyxl, json, sys, io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

EXCEL_PATH = r"C:\Users\wifik\Downloads\cerebus 3 market hoily grail (3).xlsx"
OUTPUT_PATH = r"C:\Users\wifik\Desktop\projects\larger-lab\quant-lab\holy_grail_all_data.json"

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

all_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h is not None else f"col_{j}" for j, h in enumerate(row)]
            continue
        # Skip empty rows
        non_empty = [v for v in row if v is not None]
        if not non_empty:
            continue
        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers) and val is not None:
                row_dict[headers[j]] = val
        if row_dict:
            rows.append(row_dict)
    if rows:
        all_data[sheet_name] = rows

wb.close()

# Save
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, indent=2, default=str, ensure_ascii=False)

print(f"Extracted {len(all_data)} sheets from Holy Grail Excel")
print(f"Total sheets: {len(all_data)}")
for name, rows in all_data.items():
    print(f"  {name}: {len(rows)} rows")
print(f"\nSaved to: {OUTPUT_PATH}")
